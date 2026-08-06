#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
打ち出の小づち  V805 -> V808 ビルダー（V806/V807 の内容もすべて含む）

V807 が Excel で開けなかった原因と対策
--------------------------------------
V805 のシートは「共有数式(shared formula)」を使っている。
   マスター : <c r="T3"><f t="shared" ref="T3:T66" si="4">IF(...)</f></c>
   追従     : <c r="T4"><f t="shared" si="4"/></c>
マスターのセルを書き換えると si="4" の定義が消え、追従セルが全部孤児になる。
Excel はこれを検出してシート内の数式を丸ごと削除する。
  →「削除されたレコード: /xl/worksheets/sheetN.xml パーツ内の数式」

対策：パッチを当てる前に共有数式をすべて実数式へ展開する（shared_expand.py）。
      配列数式（RSS関数など <f t="array">）には一切触らない。

V808 の内容
-----------
 A. 共有数式の展開（上記の修正）
 B. 管理シート
      M/N  損切・利確を設定セル参照＋東証の呼値へスナップ
      O    判断（決済済み行を除外、時間決済を追加）
      R    トレーリングSTOP（売買方向を判定）
      Y    重複エントリー警告
      AB   リスク均等の推奨株数
      AC   手数料・貸株料控除後の実質損益
      T/U  ★数式を外して「実約定の手入力欄」に変更
      AF/AG 設定ブロック
 C. 分析／売分析シート
      BA:BH 有効指標とゲート判定、BJ/BK ゲート設定
 D. 厳選TOP2  抽出キーを新ゲート＋新スコアへ
 E. 価格シート5枚を 250日分（E列〜IT列）に拡張
"""
import re, sys, zipfile
from datetime import datetime, date
from pathlib import Path
import openpyxl
from shared_expand import expand_file_sheets

SRC = Path(sys.argv[1] if len(sys.argv) > 1 else "v805.xlsm")
DST = Path(sys.argv[2] if len(sys.argv) > 2 else "V808.xlsm")

KANRI, GENSEN, BUNSEKI, URI = ("xl/worksheets/sheet1.xml", "xl/worksheets/sheet3.xml",
                               "xl/worksheets/sheet7.xml", "xl/worksheets/sheet8.xml")
PRICE = {"xl/worksheets/sheet11.xml": "始値", "xl/worksheets/sheet12.xml": "高値",
         "xl/worksheets/sheet13.xml": "安値", "xl/worksheets/sheet14.xml": "終値",
         "xl/worksheets/sheet15.xml": "出来高"}
HIST_DAYS, FIRST_COL = 250, 5          # E列から250日分 → 最終列 254 = IT
LAST_COL = FIRST_COL + HIST_DAYS - 1
log = []


# ------------------------------------------------------------------ XML 操作
def esc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def colname(n):
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def colnum(c):
    n = 0
    for ch in c:
        n = n * 26 + ord(ch) - 64
    return n


def _put(xml, cell, body):
    col, row = re.match(r"([A-Z]+)(\d+)", cell).groups()
    m = re.search(r'<c r="%s"[^>]*?(/>|>.*?</c>)' % cell, xml, re.S)
    if m:
        return xml[: m.start()] + body + xml[m.end():]
    rm = re.search(r'<row r="%s"[^>]*?(/>|>.*?</row>)' % row, xml, re.S)
    if not rm:
        log.append(f"  [skip] 行{row} なし ({cell})")
        return xml
    chunk = rm.group(0)
    if chunk.endswith("/>"):
        new = chunk[:-2] + ">" + body + "</row>"
    else:
        tgt = colnum(col)
        ins = next((cm.start() for cm in re.finditer(r'<c r="([A-Z]+)%s"' % row, chunk)
                    if colnum(cm.group(1)) > tgt), None)
        new = (chunk[:ins] + body + chunk[ins:]) if ins is not None else \
              (chunk[: chunk.rindex("</row>")] + body + "</row>")
    return xml[: rm.start()] + new + xml[rm.end():]


def set_f(xml, cell, f):
    """数式をセットする。既存セルの属性（書式）は保持し、キャッシュ値は捨てる。"""
    m = re.search(r'(<c r="%s"[^>]*?)(/>|>.*?</c>)' % cell, xml, re.S)
    if m:
        # 数式が変わるのでキャッシュ値は捨てるが、s=(書式) t=(型) cm= はそのまま残す。
        # 属性を落とすと Excel が「パーツ内の数式」を丸ごと削除することがある。
        return xml[: m.start()] + m.group(1) + ">" + f"<f>{esc(f)}</f>" + "</c>" + xml[m.end():]
    return _put(xml, cell, f'<c r="{cell}"><f>{esc(f)}</f></c>')


def set_t(xml, cell, text):
    return _put(xml, cell, f'<c r="{cell}" t="inlineStr"><is><t>{esc(text)}</t></is></c>')


def set_n(xml, cell, v):
    return _put(xml, cell, f'<c r="{cell}"><v>{v}</v></c>')


def freeze_value(xml, cell, v):
    """数式を外して値だけ残す。書式（日付表示など）は元のまま保持する。"""
    m = re.search(r'(<c r="%s"[^>]*?)(/>|>.*?</c>)' % cell, xml, re.S)
    if not m:
        return xml if v is None else _put(xml, cell, f'<c r="{cell}"><v>{v}</v></c>')
    attrs = re.sub(r'\s+t="[a-z]+"', "", m.group(1))   # 値型に戻すので型指定は外す
    body = "/>" if v is None else f"><v>{v}</v></c>"
    return xml[: m.start()] + attrs + body + xml[m.end():]


def rows_with(xml, col):
    return sorted(int(r) for r in re.findall(r'<c r="%s(\d+)"' % col, xml))


# ------------------------------------------------------------ 管理シート設定
# 注意：AF/AG 列は既存の「利確到達日／損切到達日」の配列数式が使用中なので使えない。
# 管理シートで完全に空いている AW/AX を使う。
SETTINGS = [("AW1", "⚙ リスク設定（AX列を変更）", None),
            ("AW2", "損切幅", 0.03),
            ("AW3", "利確幅", 0.08),
            ("AW4", "1回の許容損失額(円)", 20000),
            ("AW5", "時間決済(営業日)", 3),
            ("AW6", "片道手数料率", 0.0005),
            ("AW7", "貸株料(年率)", 0.0115)]
P_SL, P_TP, P_RISK, P_DAYS, P_FEE, P_KASHI = ("$AX$2", "$AX$3", "$AX$4",
                                              "$AX$5", "$AX$6", "$AX$7")


def tick(raw):
    """東証の呼値（TOPIX100以外の一般株）。"""
    return (f'IF({raw}<=3000,1,IF({raw}<=5000,5,IF({raw}<=30000,10,'
            f'IF({raw}<=50000,50,IF({raw}<=300000,100,500)))))')


def snap(raw, r):
    """買いは切り下げ／売りは切り上げ＝利確は約定しやすい側、損切は余裕のある側へ。
    V805 は ROUND(...,0) の1円丸めで、145件中71件(49%)が呼値に乗らず
    逆指値そのものが発注できない価格になっていた。"""
    return f'IF($C{r}="売",CEILING({raw},{tick(raw)}),FLOOR({raw},{tick(raw)}))'


def f_tp(r):
    raw = f'$F{r}*IF($C{r}="売",1-{P_TP},1+{P_TP})'
    return f'IF($F{r}="","",{snap(raw, r)})'


def f_sl(r):
    raw = f'$F{r}*IF($C{r}="売",1+{P_SL},1-{P_SL})'
    return f'IF($F{r}="","",{snap(raw, r)})'


def f_trail(r):
    """含み益が損切幅を超えるまでは固定損切をそのまま使い、超えてから建値方向へ寄せる。
    V805 は売買方向を見ずに一律「10日高値×0.97」で、買い25件中15件(60%)が
    エントリー価格より上にSTOPを置いていた。"""
    return (f'IF(OR($F{r}="",$H{r}="",$N{r}=""),"",IF($S{r}="✅売却済","",'
            f'IF($C{r}="売",'
            f'IF($H{r}>$F{r}*(1-{P_SL}),$N{r},CEILING(MIN($N{r},IF($Q{r}="",$N{r},$Q{r}*1.03)),{tick("$N"+str(r))})),'
            f'IF($H{r}<$F{r}*(1+{P_SL}),$N{r},FLOOR(MAX($N{r},IF($P{r}="",$N{r},$P{r}*0.97)),{tick("$N"+str(r))}))'
            f')))')


def f_judge(r):
    """決済済み行を本日の株価で再判定しない。V805 では利確済みの行が
    今日も「🔴損切」と表示され、ダッシュボードの件数も汚染されていた。"""
    return (f'IF(OR($B{r}="",$F{r}="",$H{r}=""),"",IF($S{r}="✅売却済","―決済済―",'
            f'IF($C{r}="売",'
            f'IF($H{r}<=$M{r},"🟢利確",IF(AND($R{r}<>"",$H{r}>=$R{r}),"🟠TRAIL決済",'
            f'IF($H{r}>=$N{r},"🔴損切",IF($W{r}>={P_DAYS},"⏱時間決済",'
            f'IF($H{r}>=$F{r}*(1+{P_SL}*0.7),"⚠️損切接近","保持継続"))))),'
            f'IF($H{r}>=$M{r},"🟢利確",IF(AND($R{r}<>"",$H{r}<=$R{r}),"🟠TRAIL決済",'
            f'IF($H{r}<=$N{r},"🔴損切",IF($W{r}>={P_DAYS},"⏱時間決済",'
            f'IF($H{r}<=$F{r}*(1-{P_SL}*0.7),"⚠️損切接近","保持継続")))))))')


def f_dup(r):
    return (f'IF($B{r}="","",IF(COUNTIFS($B$4:$B{r},$B{r},$C$4:$C{r},$C{r})>1,'
            f'"⛔重複"&COUNTIFS($B$4:$B{r},$B{r},$C$4:$C{r},$C{r})&"回目",""))')


def f_qty(r):
    return f'IF($F{r}="","",MAX(100,ROUNDDOWN({P_RISK}/($F{r}*{P_SL})/100,0)*100))'


def f_net(r):
    return (f'IF(OR($F{r}="",$G{r}="",$U{r}=""),"",ROUND($V{r}-($F{r}+$U{r})*$G{r}*{P_FEE}'
            f'-IF($C{r}="売",$F{r}*$G{r}*{P_KASHI}*$W{r}/365,0),0))')


# ------------------------------------------------- 分析／売分析シートの新指標
ACOLS = {"BA": "14日平均レンジ率(%)", "BB": "25日レンジ幅(%)", "BC": "ギャップ率(%)",
         "BD": "25日高値距離(%)", "BE": "日中レンジ率(%)", "BF": "V808スコア",
         "BG": "ゲート通過", "BH": "判定"}
GATE = [("BJ1", "⚙ 抽出ゲート（BK列を変更）", None),
        ("BJ2", "14日平均レンジ率 上限(%)", 2.2),
        ("BJ3", "25日レンジ幅 上限(%)", 9.0),
        ("BJ4", "ギャップ率 上限(%)", 0.5),
        ("BJ5", "前日比 上限(%)", 1.5),
        ("BJ6", "前日比 下限(%)", -3.0),
        ("BJ7", "25日高値距離 下限(%)", -10.0),
        ("BJ8", "日中レンジ率 上限(%)", 2.0),
        ("BJ9", "1日の採用上限(件)", 3),
        ("BJ10", "この方向を使う 1=使う 0=停止", 1)]
K = {n: f"$BK${n}" for n in range(2, 11)}
AROW_FIRST, AROW_LAST = 3, 206


def midx(r):
    return (f'IFERROR(MATCH($C{r},銘柄管理!$B$6:$B$305,0),'
            f'MATCH(TEXT($C{r},"0"),銘柄管理!$B$6:$B$305,0))-1')


def off(sheet, r, n):
    return f'OFFSET({sheet}!$E$6,{midx(r)},0,1,{n})'


def z(col, r):
    rng = f'${col}${AROW_FIRST}:${col}${AROW_LAST}'
    return f'(${col}{r}-AVERAGE({rng}))/STDEV({rng})'


def patch_analysis(xml, enable, title):
    for c, h in ACOLS.items():
        xml = set_t(xml, f"{c}2", h)
    for r in range(AROW_FIRST, AROW_LAST + 1):
        # 欠損日があっても壊れないよう、配列除算を使わない加重平均形にする
        # （真の14日平均レンジ率との相関 0.9989）
        xml = set_f(xml, f"BA{r}",
                    f'IFERROR((SUM({off("高値", r, 14)})-SUM({off("安値", r, 14)}))'
                    f'/SUM({off("終値", r, 14)})*100,"")')
        xml = set_f(xml, f"BB{r}",
                    f'IFERROR((MAX({off("高値", r, 25)})-MIN({off("安値", r, 25)}))/$E{r}*100,"")')
        xml = set_f(xml, f"BC{r}", f'IFERROR(($F{r}-$G{r})/$G{r}*100,"")')
        xml = set_f(xml, f"BD{r}",
                    f'IFERROR(($E{r}-MAX({off("高値", r, 25)}))/MAX({off("高値", r, 25)})*100,"")')
        xml = set_f(xml, f"BE{r}", f'IFERROR(($H{r}-$I{r})/$F{r}*100,"")')
        xml = set_f(xml, f"BF{r}",
                    'IFERROR(-(' + z("BA", r) + ')-(' + z("BE", r) + ')-(' + z("BB", r) +
                    ')-(' + z("Z", r) + ')-(' + z("BC", r) + '),"")')
        xml = set_f(xml, f"BG{r}",
                    f'IF(OR($C{r}="",$C{r}="TOPX",$BA{r}="",$BF{r}=""),0,'
                    f'IF(AND({K[10]}=1,$BA{r}<={K[2]},$BB{r}<={K[3]},$BC{r}<={K[4]},'
                    f'$Z{r}<={K[5]},$Z{r}>={K[6]},$BD{r}>={K[7]},$BE{r}<={K[8]}),1,0))')
        xml = set_f(xml, f"BH{r}",
                    f'IF($C{r}="","",IF($BG{r}=1,"○ゲート通過",'
                    f'IF({K[10]}<>1,"― この方向は停止中",'
                    f'IF($BA{r}>{K[2]},"×値動きが荒い",IF($BB{r}>{K[3]},"×レンジが広い",'
                    f'IF($BC{r}>{K[4]},"×窓を開けて上昇",IF($Z{r}>{K[5]},"×前日に急騰",'
                    f'IF($Z{r}<{K[6]},"×前日に急落",IF($BD{r}<{K[7]},"×高値から離れすぎ",'
                    f'IF($BE{r}>{K[8]},"×当日の値幅が大きい","×")))))))))')
    for cell, label, val in GATE:
        xml = set_t(xml, cell, label)
        if val is not None:
            xml = set_n(xml, "BK" + cell[2:], enable if cell == "BJ10" else val)
    log.append(f"{title}: 指標8列×{AROW_LAST - AROW_FIRST + 1}行 と ゲート設定 BJ1:BK10 を追加")
    return xml


def widen_refs(xml, name):
    """数式内にハードコードされた「価格シートの参照範囲の右端」を250日分へ広げる。
    管理シートは $BT$（68日）、分析／売分析は $CZ$（100日）で止まっていた。
    ここを広げないと、価格シートを250日に増やしても数式が古い範囲しか見ない。"""
    end = colname(LAST_COL)
    pat = re.compile(r'(始値|高値|安値|終値|出来高)!(\$[A-Z]+\$\d+):\$(?:BT|CZ)\$(\d+)')
    xml, n = pat.subn(lambda m: f'{m.group(1)}!{m.group(2)}:${end}${m.group(3)}', xml)
    if n:
        log.append(f"  {name}: 価格シート参照の右端を ${end}$ へ拡張（{n}箇所）")
    return xml


def fix_geometry(xml, name):
    """行の spans 属性と worksheet の dimension を、実際のセルに合わせて直す。

    spans は「その行に存在する列の範囲」の宣言。新しい列を足したのに
    宣言を直さないと実態とずれ、Excel が数式を丸ごと落とすことがある。
    dimension も同様に、シート全体の使用範囲を宣言し直す。"""
    lo_all, hi_all, row_min, row_max = 10 ** 9, 0, 10 ** 9, 0

    def fix_row(m):
        nonlocal lo_all, hi_all, row_min, row_max
        r, attrs, body = int(m.group(1)), m.group(2), m.group(3)
        cols = [colnum(c) for c in re.findall(r'<c r="([A-Z]+)%d"' % r, m.group(0))]
        if not cols:
            return m.group(0)
        lo, hi = min(cols), max(cols)
        lo_all, hi_all = min(lo_all, lo), max(hi_all, hi)
        row_min, row_max = min(row_min, r), max(row_max, r)
        if 'spans="' in attrs:
            attrs = re.sub(r'spans="\d+:\d+"', f'spans="{lo}:{hi}"', attrs)
        else:
            attrs = f' spans="{lo}:{hi}"' + attrs
        return f'<row r="{r}"{attrs}{body}'

    xml = re.sub(r'<row r="(\d+)"([^>]*?)((?:/>|>(?:(?!</row>).)*</row>))',
                 fix_row, xml, flags=re.S)
    if hi_all:
        ref = f"{colname(lo_all)}{row_min}:{colname(hi_all)}{row_max}"
        xml = re.sub(r'<dimension ref="[^"]+"/>', f'<dimension ref="{ref}"/>', xml)
        log.append(f"  {name}: spans と dimension を実態に合わせて再計算 → {ref}")
    return xml


# ---------------------------------------------------------- 価格シートの拡張
def extend_price(xml, name):
    xml = re.sub(r'<dimension ref="A1:[A-Z]+(\d+)"/>',
                 lambda m: f'<dimension ref="A1:{colname(LAST_COL)}{m.group(1)}"/>', xml)
    if "<cols>" in xml:
        xml = xml.replace("</cols>",
                          f'<col min="105" max="{LAST_COL}" width="7" style="1" '
                          f'customWidth="1"/></cols>')
    xml = re.sub(r'spans="1:\d+"', f'spans="1:{LAST_COL}"', xml)
    log.append(f"  {name}: E列〜{colname(LAST_COL)}列（{HIST_DAYS}営業日分）に拡張")
    return xml


# -------------------------------------------------------------------- 実行
def main():
    zin = zipfile.ZipFile(SRC)
    items = {n: zin.read(n) for n in zin.namelist()}
    order = zin.namelist()
    zin.close()

    log.append("【A】共有数式の展開（V807 が開けなかった原因の修正）")
    log.extend(expand_file_sheets(SRC, items, {
        KANRI: "管理", GENSEN: "厳選TOP2", BUNSEKI: "分析", URI: "売分析"}))

    # --- B. 管理シート ---
    wb_v = openpyxl.load_workbook(SRC, data_only=True)
    kv = wb_v["管理"]
    x = items[KANRI].decode()
    trade_rows = [r for r in rows_with(x, "N") if r >= 4]
    n_frozen = 0
    for r in trade_rows:
        x = set_f(x, f"M{r}", f_tp(r))
        x = set_f(x, f"N{r}", f_sl(r))
        x = set_f(x, f"R{r}", f_trail(r))
        x = set_f(x, f"O{r}", f_judge(r))
        x = set_f(x, f"Y{r}", f_dup(r))
        x = set_f(x, f"AB{r}", f_qty(r))
        x = set_f(x, f"AC{r}", f_net(r))
        # ★ T列(売却日)・U列(売却価格) を実約定の手入力欄にする。
        #   V805 では両方とも数式で、売却価格は必ず利確ラインか損切ライン
        #   ちょうどを返していた（スリッページもギャップも0の仮定値）。
        #   既に決済済みの行は現在の表示値を静的な値として凍結し、
        #   保有中の行は空欄にする。以後は実際の約定を手で入れる。
        tv, uv = kv.cell(r, 20).value, kv.cell(r, 21).value
        if isinstance(tv, datetime):
            tv = tv.date()
        if isinstance(tv, date) and isinstance(uv, (int, float)):
            x = freeze_value(x, f"T{r}", (tv - date(1899, 12, 30)).days)
            x = freeze_value(x, f"U{r}", uv)
            n_frozen += 1
        else:
            x = freeze_value(x, f"T{r}", None)
            x = freeze_value(x, f"U{r}", None)
    for cell, label, val in SETTINGS:
        x = set_t(x, cell, label)
        if val is not None:
            x = set_n(x, "AX" + cell[2:], val)
    for cell, txt in [("Y3", "重複警告"), ("AB3", "推奨株数"),
                      ("AC3", "実質損益(手数料後)"),
                      ("T3", "売却日★実約定を手入力"),
                      ("U3", "売却価格★実約定を手入力")]:
        x = set_t(x, cell, txt)
    items[KANRI] = x.encode()
    log.append(f"\n【B】管理シート: {len(trade_rows)}行に適用")
    log.append("  M/N 損切・利確を設定セル参照＋東証の呼値へスナップ（損切3%・利確8%）")
    log.append("  R   トレーリングSTOP に売買方向判定を追加")
    log.append("  O   判断から決済済み行を除外、⏱時間決済を追加")
    log.append("  Y/AB/AC 重複警告・推奨株数・実質損益 を新設")
    log.append(f"  ★T/U 数式を削除して実約定の手入力欄に変更"
               f"（決済済み{n_frozen}行は現在値を凍結、保有中は空欄）")

    # --- C. 分析／売分析 ---
    items[BUNSEKI] = patch_analysis(items[BUNSEKI].decode(), 1, "\n【C】分析シート（買い）").encode()
    items[URI] = patch_analysis(items[URI].decode(), 0, "売分析シート（売り・既定で停止）").encode()

    # --- D. 厳選TOP2 ---
    x = items[GENSEN].decode()
    for r in range(AROW_FIRST, AROW_LAST + 1):
        x = set_f(x, f"T{r}", f'IF(AND(分析!$C{r}<>"",分析!$C{r}<>"TOPX",N(分析!$BG{r})=1),'
                              f'(N(分析!$BF{r})+100)*10000-ROW(),"")')
        x = set_f(x, f"V{r}", f'IF(AND(売分析!$C{r}<>"",売分析!$C{r}<>"TOPX",N(売分析!$BG{r})=1),'
                              f'(N(売分析!$BF{r})+100)*10000-ROW(),"")')
    x = set_n(x, "B5", 3)
    for cell, txt in [("A6", "【買い】ゲート設定は 分析シート BJ1:BK10"),
                      ("A7", "旧スコアしきい値（現在は未使用）"),
                      ("A10", "【売り】既定で停止 ― 売分析!BK10 を 1 で有効化"),
                      ("A11", "旧スコアしきい値（現在は未使用）"),
                      ("A14", "※ 順位付け：新スコアの高い順。ゲート全滅の日は候補0件＝見送りが正常です。")]:
        x = set_t(x, cell, txt)
    items[GENSEN] = x.encode()
    log.append("\n【D】厳選TOP2: 抽出キー T列/V列 を新ゲート＋新スコアへ、採用上限3件")

    # --- E. 価格シートを250日に拡張 ---
    log.append(f"\n【E】価格シートを{HIST_DAYS}営業日分に拡張")
    # 管理シートは対象外。AE〜AI の到達日シミュレーション列は 3,664 個の配列数式で、
    # T/U を手入力にした今は使われない。触らないことでリスクを減らす。
    for path, name in ((BUNSEKI, "分析"), (URI, "売分析")):
        items[path] = widen_refs(items[path].decode(), name).encode()
    for path, name in PRICE.items():
        items[path] = extend_price(items[path].decode(), name).encode()

    # --- 行の spans と dimension を実態に合わせる ---
    log.append("\n【F】幾何情報の整合")
    for path, name in ((KANRI, "管理"), (GENSEN, "厳選TOP2"),
                       (BUNSEKI, "分析"), (URI, "売分析")):
        items[path] = fix_geometry(items[path].decode(), name).encode()

    # --- 仕上げ ---
    if "xl/calcChain.xml" in items:
        items["[Content_Types].xml"] = re.sub(
            r'<Override PartName="/xl/calcChain\.xml"[^>]*/>', "",
            items["[Content_Types].xml"].decode()).encode()
        items["xl/_rels/workbook.xml.rels"] = re.sub(
            r'<Relationship[^>]*Target="calcChain\.xml"[^>]*/>', "",
            items["xl/_rels/workbook.xml.rels"].decode()).encode()
    wbx = items["xl/workbook.xml"].decode()
    wbx = (re.sub(r"<calcPr[^>]*/>", '<calcPr calcId="0" fullCalcOnLoad="1"/>', wbx)
           if "<calcPr" in wbx else
           wbx.replace("</workbook>", '<calcPr calcId="0" fullCalcOnLoad="1"/></workbook>'))
    items["xl/workbook.xml"] = wbx.encode()

    with zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED) as zo:
        for n in order:
            if n == "xl/calcChain.xml":
                continue
            zo.writestr(n, items[n])
    log.append(f"\n出力: {DST}  ({DST.stat().st_size:,} bytes)")
    print("\n".join(log))


if __name__ == "__main__":
    main()
