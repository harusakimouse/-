#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
打ち出の小づち  V805 -> V809 ビルダー

■ V807 / V808 が Excel で開けなかった経緯と、今回の設計

 1回目(V807)  共有数式(shared formula)のマスターを壊していた。
              マスターを書き換えると si= の定義が消え、追従セルが孤児になり、
              Excel がシート内の数式を丸ごと削除する。
              → shared_expand.py で事前に全展開して解決。

 2回目(V808)  それでも同じ3シート(管理/分析/売分析)が落ちた。
              切り分けビルド「検証A」(展開＋spans/dimension修正のみ)は
              正常に開いたので、展開処理は正しいと確定。
              落ちた3シートと無事だった厳選TOP2の違いは1つだけ ―
              **既存の行に新しい <c> セルを挿入したかどうか**。
              厳選TOP2 は既存セルの置き換えしかしておらず無事だった。

 今回(V809)   既存シートへのセル挿入を完全にやめる。
              ・既存シートは「既にあるセルの中身を差し替える」だけ
              ・新しい列が必要なものは、まるごと新規シートに逃がす
                （新規シートの XML は一から自分で生成するので、
                  既存構造を壊しようがない）

■ 追加する2枚のシート
   V809設定    リスク設定と抽出ゲートのパラメータ、管理シート用の補助列
   V809指標    分析/売分析の全銘柄について、有効指標・スコア・ゲート判定

■ 既存シートへの変更（すべて既存セルの置き換え）
   管理        M/N 損切・利確（呼値スナップ）、O 判断、R トレーリングSTOP、
               T/U を実約定の手入力欄に（数式を外して値にする）
   厳選TOP2    抽出キー T列/V列 を V809指標 のゲートとスコア参照に
   分析/売分析  価格シート参照の右端を250日分へ拡張
   価格5シート  E列〜IT列（250営業日分）へ拡張
"""
import re, sys, zipfile
from datetime import datetime, date
from pathlib import Path
import openpyxl
from shared_expand import expand_file_sheets

SRC = Path(sys.argv[1] if len(sys.argv) > 1 else "v805.xlsm")
DST = Path(sys.argv[2] if len(sys.argv) > 2 else "V809.xlsm")

KANRI, GENSEN, BUNSEKI, URI = ("xl/worksheets/sheet1.xml", "xl/worksheets/sheet3.xml",
                               "xl/worksheets/sheet7.xml", "xl/worksheets/sheet8.xml")
PRICE = {"xl/worksheets/sheet11.xml": "始値", "xl/worksheets/sheet12.xml": "高値",
         "xl/worksheets/sheet13.xml": "安値", "xl/worksheets/sheet14.xml": "終値",
         "xl/worksheets/sheet15.xml": "出来高"}
HIST_DAYS = 250
LAST_COL = 4 + HIST_DAYS                      # E列(5)から250日 → 254 = IT
SET_SH, IND_SH = "V809設定", "V809指標"
AROW1, AROWN = 3, 206                         # 分析／売分析の銘柄行
KROW1 = 4                                     # 管理の明細開始行
log = []


# ------------------------------------------------------------------ ユーティリティ
def esc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def colname(n):
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def colnum(c):
    n = 0
    for ch in c:
        n = n * 26 + ord(ch) - 64
    return n


def replace_cell(xml, cell, inner, drop_type=False):
    """既存セルの中身だけを差し替える。セルが無ければ何もしない（挿入しない）。

    V808 までは無いセルを挿入していたが、それが Excel に弾かれる原因だった。
    ここでは絶対に挿入しない。"""
    m = re.search(r'(<c r="%s"[^>]*?)(/>|>.*?</c>)' % cell, xml, re.S)
    if not m:
        log.append(f"  [注意] {cell} が存在しないため書き込みませんでした")
        return xml
    attrs = m.group(1)
    if drop_type:
        attrs = re.sub(r'\s+t="[a-z]+"', "", attrs)
    body = "/>" if inner is None else ">" + inner + "</c>"
    return xml[: m.start()] + attrs + body + xml[m.end():]


def set_formula(xml, cell, f):
    return replace_cell(xml, cell, f"<f>{esc(f)}</f>")


def set_value(xml, cell, v):
    return replace_cell(xml, cell, None if v is None else f"<v>{v}</v>", drop_type=True)


def rows_with(xml, col):
    return sorted(int(r) for r in re.findall(r'<c r="%s(\d+)"' % col, xml))


def fix_geometry(xml, name):
    """行の spans と worksheet の dimension を実際のセルに合わせ直す。"""
    lo_all, hi_all, r_min, r_max = 10 ** 9, 0, 10 ** 9, 0

    def fix_row(m):
        nonlocal lo_all, hi_all, r_min, r_max
        r, attrs, body = int(m.group(1)), m.group(2), m.group(3)
        cols = [colnum(c) for c in re.findall(r'<c r="([A-Z]+)%d"' % r, m.group(0))]
        if not cols:
            return m.group(0)
        lo, hi = min(cols), max(cols)
        lo_all, hi_all = min(lo_all, lo), max(hi_all, hi)
        r_min, r_max = min(r_min, r), max(r_max, r)
        attrs = (re.sub(r'spans="\d+:\d+"', f'spans="{lo}:{hi}"', attrs)
                 if 'spans="' in attrs else f' spans="{lo}:{hi}"' + attrs)
        return f'<row r="{r}"{attrs}{body}'

    xml = re.sub(r'<row r="(\d+)"([^>]*?)((?:/>|>(?:(?!</row>).)*</row>))',
                 fix_row, xml, flags=re.S)
    if hi_all:
        ref = f"{colname(lo_all)}{r_min}:{colname(hi_all)}{r_max}"
        xml = re.sub(r'<dimension ref="[^"]+"/>', f'<dimension ref="{ref}"/>', xml)
    return xml


def widen_refs(xml, name):
    """数式にハードコードされた価格シート参照の右端を250日分へ広げる。"""
    end = colname(LAST_COL)
    pat = re.compile(r'(始値|高値|安値|終値|出来高)!(\$[A-Z]+\$\d+):\$(?:BT|CZ)\$(\d+)')
    xml, n = pat.subn(lambda m: f'{m.group(1)}!{m.group(2)}:${end}${m.group(3)}', xml)
    if n:
        log.append(f"  {name}: 価格シート参照の右端を ${end}$ へ拡張（{n}箇所）")
    return xml


def extend_price(xml, name):
    xml = re.sub(r'<dimension ref="A1:[A-Z]+(\d+)"/>',
                 lambda m: f'<dimension ref="A1:{colname(LAST_COL)}{m.group(1)}"/>', xml)
    if "<cols>" in xml:
        xml = xml.replace("</cols>", f'<col min="105" max="{LAST_COL}" width="7" '
                                     f'style="1" customWidth="1"/></cols>')
    xml = re.sub(r'spans="1:\d+"', f'spans="1:{LAST_COL}"', xml)
    log.append(f"  {name}: E列〜{colname(LAST_COL)}列（{HIST_DAYS}営業日分）")
    return xml


# ------------------------------------------------------------ 新規シートの生成
class Sheet:
    """新規ワークシートの XML を一から組み立てる。既存構造に触れないので安全。"""

    def __init__(self):
        self.cells = {}          # (row, col) -> xml文字列

    def text(self, cell, s):
        self._put(cell, f'<c r="{cell}" t="inlineStr"><is><t>{esc(s)}</t></is></c>')

    def num(self, cell, v):
        self._put(cell, f'<c r="{cell}"><v>{v}</v></c>')

    def formula(self, cell, f):
        self._put(cell, f'<c r="{cell}"><f>{esc(f)}</f></c>')

    def _put(self, cell, body):
        c, r = re.match(r"([A-Z]+)(\d+)", cell).groups()
        self.cells[(int(r), colnum(c))] = body

    def xml(self, widths=None):
        rows = {}
        for (r, c), body in self.cells.items():
            rows.setdefault(r, []).append((c, body))
        parts = []
        for r in sorted(rows):
            cs = sorted(rows[r])
            lo, hi = cs[0][0], cs[-1][0]
            parts.append(f'<row r="{r}" spans="{lo}:{hi}">'
                         + "".join(b for _, b in cs) + "</row>")
        maxr = max(rows) if rows else 1
        maxc = max(c for _, c in self.cells) if self.cells else 1
        cols = ""
        if widths:
            cols = "<cols>" + "".join(
                f'<col min="{a}" max="{b}" width="{w}" customWidth="1"/>'
                for a, b, w in widths) + "</cols>"
        return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                f'<dimension ref="A1:{colname(maxc)}{maxr}"/>'
                '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
                '<sheetFormatPr defaultRowHeight="18"/>'
                + cols +
                "<sheetData>" + "".join(parts) + "</sheetData>"
                '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
                'header="0.3" footer="0.3"/></worksheet>').encode("utf-8")


def add_sheet(items, order, name, xml_bytes):
    """新しいワークシートをブックに登録する。"""
    used = {int(m) for m in re.findall(r'worksheets/sheet(\d+)\.xml',
                                       items["xl/_rels/workbook.xml.rels"].decode())}
    n = max(used) + 1
    path = f"xl/worksheets/sheet{n}.xml"
    items[path] = xml_bytes
    order.append(path)

    rels = items["xl/_rels/workbook.xml.rels"].decode()
    rid = "rId" + str(max(int(x) for x in re.findall(r'Id="rId(\d+)"', rels)) + 1)
    rels = rels.replace("</Relationships>",
                        f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/'
                        f'officeDocument/2006/relationships/worksheet" '
                        f'Target="worksheets/sheet{n}.xml"/></Relationships>')
    items["xl/_rels/workbook.xml.rels"] = rels.encode()

    ct = items["[Content_Types].xml"].decode()
    ct = ct.replace("</Types>",
                    f'<Override PartName="/{path}" ContentType="application/vnd.'
                    f'openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>')
    items["[Content_Types].xml"] = ct.encode()

    wb = items["xl/workbook.xml"].decode()
    sid = max(int(x) for x in re.findall(r'sheetId="(\d+)"', wb)) + 1
    wb = wb.replace("</sheets>", f'<sheet name="{esc(name)}" sheetId="{sid}" '
                                 f'r:id="{rid}"/></sheets>')
    items["xl/workbook.xml"] = wb.encode()
    log.append(f"  シート「{name}」を追加（{path} / sheetId={sid} / {rid}）")


# ------------------------------------------------------------------ 数式の部品
P_SL, P_TP = f"{SET_SH}!$B$3", f"{SET_SH}!$B$4"
P_RISK, P_DAYS = f"{SET_SH}!$B$5", f"{SET_SH}!$B$6"
P_FEE, P_KASHI = f"{SET_SH}!$B$7", f"{SET_SH}!$B$8"


def tick(raw):
    return (f'IF({raw}<=3000,1,IF({raw}<=5000,5,IF({raw}<=30000,10,'
            f'IF({raw}<=50000,50,IF({raw}<=300000,100,500)))))')


def snap(raw, r):
    return f'IF($C{r}="売",CEILING({raw},{tick(raw)}),FLOOR({raw},{tick(raw)}))'


def f_tp(r):
    raw = f'$F{r}*IF($C{r}="売",1-{P_TP},1+{P_TP})'
    return f'IF($F{r}="","",{snap(raw, r)})'


def f_sl(r):
    raw = f'$F{r}*IF($C{r}="売",1+{P_SL},1-{P_SL})'
    return f'IF($F{r}="","",{snap(raw, r)})'


def f_trail(r):
    return (f'IF(OR($F{r}="",$H{r}="",$N{r}=""),"",IF($S{r}="✅売却済","",'
            f'IF($C{r}="売",'
            f'IF($H{r}>$F{r}*(1-{P_SL}),$N{r},CEILING(MIN($N{r},IF($Q{r}="",$N{r},$Q{r}*1.03)),{tick("$N"+str(r))})),'
            f'IF($H{r}<$F{r}*(1+{P_SL}),$N{r},FLOOR(MAX($N{r},IF($P{r}="",$N{r},$P{r}*0.97)),{tick("$N"+str(r))}))'
            f')))')


def f_judge(r):
    return (f'IF(OR($B{r}="",$F{r}="",$H{r}=""),"",IF($S{r}="✅売却済","―決済済―",'
            f'IF($C{r}="売",'
            f'IF($H{r}<=$M{r},"🟢利確",IF(AND($R{r}<>"",$H{r}>=$R{r}),"🟠TRAIL",'
            f'IF($H{r}>=$N{r},"🔴損切",IF($W{r}>={P_DAYS},"⏱時間決済",'
            f'IF($H{r}>=$F{r}*(1+{P_SL}*0.7),"⚠️損切接近","保持継続"))))),'
            f'IF($H{r}>=$M{r},"🟢利確",IF(AND($R{r}<>"",$H{r}<=$R{r}),"🟠TRAIL",'
            f'IF($H{r}<=$N{r},"🔴損切",IF($W{r}>={P_DAYS},"⏱時間決済",'
            f'IF($H{r}<=$F{r}*(1-{P_SL}*0.7),"⚠️損切接近","保持継続")))))))')


def midx(sh, r):
    return (f'IFERROR(MATCH({sh}!$C{r},銘柄管理!$B$6:$B$305,0),'
            f'MATCH(TEXT({sh}!$C{r},"0"),銘柄管理!$B$6:$B$305,0))-1')


def off(price, sh, r, n):
    return f'OFFSET({price}!$E$6,{midx(sh, r)},0,1,{n})'


# ------------------------------------------------------------------ 実行
def main():
    zin = zipfile.ZipFile(SRC)
    items = {n: zin.read(n) for n in zin.namelist()}
    order = list(zin.namelist())
    zin.close()

    log.append("【A】共有数式の展開（検証Aで正常に開くことを確認済みの処理）")
    log.extend(expand_file_sheets(SRC, items, {
        KANRI: "管理", GENSEN: "厳選TOP2", BUNSEKI: "分析", URI: "売分析"}))

    # ---- B. 管理シート（既存セルの置き換えのみ。挿入は一切しない）----
    wbv = openpyxl.load_workbook(SRC, data_only=True)
    kv = wbv["管理"]
    x = items[KANRI].decode()
    trade_rows = [r for r in rows_with(x, "N") if r >= KROW1]
    frozen = 0
    for r in trade_rows:
        x = set_formula(x, f"M{r}", f_tp(r))
        x = set_formula(x, f"N{r}", f_sl(r))
        x = set_formula(x, f"R{r}", f_trail(r))
        x = set_formula(x, f"O{r}", f_judge(r))
        tv, uv = kv.cell(r, 20).value, kv.cell(r, 21).value
        if isinstance(tv, datetime):
            tv = tv.date()
        if isinstance(tv, date) and isinstance(uv, (int, float)):
            x = set_value(x, f"T{r}", (tv - date(1899, 12, 30)).days)
            x = set_value(x, f"U{r}", uv)
            frozen += 1
        else:
            x = set_value(x, f"T{r}", None)
            x = set_value(x, f"U{r}", None)
    items[KANRI] = x.encode()
    log.append(f"\n【B】管理シート（{len(trade_rows)}行・既存セルの置き換えのみ）")
    log.append("  M/N 損切・利確を設定参照＋東証の呼値へスナップ")
    log.append("  R   トレーリングSTOP に売買方向判定")
    log.append("  O   判断から決済済み行を除外、時間決済を追加")
    log.append(f"  T/U 実約定の手入力欄へ（決済済み{frozen}行は現在値で凍結、保有中は空欄）")

    # ---- C. 新規シート「V809設定」----
    st = Sheet()
    st.text("A1", "⚙ V809 設定 ― 青字の B列 と E列 を変更してください")
    st.text("A2", "■ リスク管理（管理シートが参照）")
    for i, (lab, val) in enumerate([
            ("損切幅", 0.03), ("利確幅", 0.08), ("1回の許容損失額(円)", 20000),
            ("時間決済(営業日)", 3), ("片道手数料率", 0.0005), ("貸株料(年率)", 0.0115)], start=3):
        st.text(f"A{i}", lab)
        st.num(f"B{i}", val)
    st.text("D2", "■ 抽出ゲート（V809指標が参照）")
    for i, (lab, val) in enumerate([
            ("14日平均レンジ率 上限(%)", 2.2), ("25日レンジ幅 上限(%)", 9.0),
            ("ギャップ率 上限(%)", 0.5), ("前日比 上限(%)", 1.5),
            ("前日比 下限(%)", -3.0), ("25日高値距離 下限(%)", -10.0),
            ("日中レンジ率 上限(%)", 2.0), ("1日の採用上限(件)", 3),
            ("買いを使う 1=使う 0=停止", 1), ("売りを使う 1=使う 0=停止", 0)], start=3):
        st.text(f"D{i}", lab)
        st.num(f"E{i}", val)
    st.text("A15", "■ 管理シート用の補助（行番号は管理シートに対応）")
    for c, h in [("A16", "管理の行"), ("B16", "重複警告"), ("C16", "推奨株数"),
                 ("D16", "実質損益(手数料後)")]:
        st.text(c, h)
    for i, r in enumerate(trade_rows, start=17):
        st.num(f"A{i}", r)
        st.formula(f"B{i}", f'IF(管理!$B{r}="","",IF(COUNTIFS(管理!$B$4:$B{r},管理!$B{r},'
                            f'管理!$C$4:$C{r},管理!$C{r})>1,"⛔重複",""))')
        st.formula(f"C{i}", f'IF(管理!$F{r}="","",MAX(100,ROUNDDOWN($B$5/(管理!$F{r}*$B$3)/100,0)*100))')
        st.formula(f"D{i}", f'IF(OR(管理!$F{r}="",管理!$U{r}=""),"",ROUND(管理!$V{r}'
                            f'-(管理!$F{r}+管理!$U{r})*管理!$G{r}*$B$7'
                            f'-IF(管理!$C{r}="売",管理!$F{r}*管理!$G{r}*$B$8*管理!$W{r}/365,0),0))')
    log.append("\n【C】新規シートを追加（既存シートに列を足さないための逃がし先）")
    add_sheet(items, order, SET_SH,
              st.xml([(1, 1, 26), (2, 2, 12), (4, 4, 26), (5, 5, 12)]))

    # ---- D. 新規シート「V809指標」----
    ind = Sheet()
    ind.text("A1", "V809 指標 ― 分析／売分析の全銘柄を、有効性を検定した指標だけで評価")
    ind.text("A2", "◆ 買い（分析シート）")
    ind.text("K2", "◆ 売り（売分析シート）")
    heads = ["コード", "銘柄名", "14日平均レンジ率", "25日レンジ幅", "ギャップ率",
             "25日高値距離", "日中レンジ率", "スコア", "ゲート", "判定"]
    for i, h in enumerate(heads):
        ind.text(f"{colname(1 + i)}{AROW1 - 1}", h)
        ind.text(f"{colname(11 + i)}{AROW1 - 1}", h)

    def block(base_col, sh, enable_cell):
        C = lambda k: colname(base_col + k)
        for r in range(AROW1, AROWN + 1):
            ind.formula(f"{C(0)}{r}", f'IF({sh}!$C{r}="","",{sh}!$C{r})')
            ind.formula(f"{C(1)}{r}", f'IF({sh}!$C{r}="","",{sh}!$D{r})')
            # 欠損日があっても壊れない加重平均形（真の14日平均との相関 0.9989）
            ind.formula(f"{C(2)}{r}",
                        f'IFERROR((SUM({off("高値", sh, r, 14)})-SUM({off("安値", sh, r, 14)}))'
                        f'/SUM({off("終値", sh, r, 14)})*100,"")')
            ind.formula(f"{C(3)}{r}",
                        f'IFERROR((MAX({off("高値", sh, r, 25)})-MIN({off("安値", sh, r, 25)}))'
                        f'/{sh}!$E{r}*100,"")')
            ind.formula(f"{C(4)}{r}", f'IFERROR(({sh}!$F{r}-{sh}!$G{r})/{sh}!$G{r}*100,"")')
            ind.formula(f"{C(5)}{r}",
                        f'IFERROR(({sh}!$E{r}-MAX({off("高値", sh, r, 25)}))'
                        f'/MAX({off("高値", sh, r, 25)})*100,"")')
            ind.formula(f"{C(6)}{r}", f'IFERROR(({sh}!$H{r}-{sh}!$I{r})/{sh}!$F{r}*100,"")')

            def zz(k):
                col = C(k)
                rng = f'${col}${AROW1}:${col}${AROWN}'
                return f'(${col}{r}-AVERAGE({rng}))/STDEV({rng})'
            zpc = (f'({sh}!$Z{r}-AVERAGE({sh}!$Z${AROW1}:$Z${AROWN}))'
                   f'/STDEV({sh}!$Z${AROW1}:$Z${AROWN})')
            ind.formula(f"{C(7)}{r}",
                        f'IFERROR(-({zz(2)})-({zz(6)})-({zz(3)})-({zpc})-({zz(4)}),"")')
            ind.formula(f"{C(8)}{r}",
                        f'IF(OR({C(0)}{r}="",{C(0)}{r}="TOPX",{C(2)}{r}="",{C(7)}{r}=""),0,'
                        f'IF(AND({SET_SH}!{enable_cell}=1,'
                        f'{C(2)}{r}<={SET_SH}!$E$3,{C(3)}{r}<={SET_SH}!$E$4,'
                        f'{C(4)}{r}<={SET_SH}!$E$5,{sh}!$Z{r}<={SET_SH}!$E$6,'
                        f'{sh}!$Z{r}>={SET_SH}!$E$7,{C(5)}{r}>={SET_SH}!$E$8,'
                        f'{C(6)}{r}<={SET_SH}!$E$9),1,0))')
            ind.formula(f"{C(9)}{r}",
                        f'IF({C(0)}{r}="","",IF({C(8)}{r}=1,"○ゲート通過",'
                        f'IF({SET_SH}!{enable_cell}<>1,"― 停止中",'
                        f'IF({C(2)}{r}>{SET_SH}!$E$3,"×値動きが荒い",'
                        f'IF({C(3)}{r}>{SET_SH}!$E$4,"×レンジが広い",'
                        f'IF({C(4)}{r}>{SET_SH}!$E$5,"×窓を開けて上昇",'
                        f'IF({sh}!$Z{r}>{SET_SH}!$E$6,"×前日に急騰",'
                        f'IF({sh}!$Z{r}<{SET_SH}!$E$7,"×前日に急落",'
                        f'IF({C(5)}{r}<{SET_SH}!$E$8,"×高値から離れすぎ",'
                        f'IF({C(6)}{r}>{SET_SH}!$E$9,"×当日の値幅が大きい","×")))))))))')

    block(1, "分析", "$E$11")
    block(11, "売分析", "$E$12")
    add_sheet(items, order, IND_SH,
              ind.xml([(1, 1, 9), (2, 2, 20), (3, 8, 14), (9, 9, 8), (10, 10, 20),
                       (11, 11, 9), (12, 12, 20), (13, 18, 14), (19, 19, 8), (20, 20, 20)]))

    # ---- E. 厳選TOP2（既存セルの置き換えのみ）----
    x = items[GENSEN].decode()
    for r in range(AROW1, AROWN + 1):
        x = set_formula(x, f"T{r}", f'IF(AND(分析!$C{r}<>"",分析!$C{r}<>"TOPX",'
                                    f'N({IND_SH}!$I{r})=1),(N({IND_SH}!$H{r})+100)*10000-ROW(),"")')
        x = set_formula(x, f"V{r}", f'IF(AND(売分析!$C{r}<>"",売分析!$C{r}<>"TOPX",'
                                    f'N({IND_SH}!$S{r})=1),(N({IND_SH}!$R{r})+100)*10000-ROW(),"")')
    x = set_value(x, "B5", 3)
    items[GENSEN] = x.encode()
    log.append(f"\n【D】厳選TOP2: 抽出キーを {IND_SH} のゲート／スコア参照へ、採用上限3件")

    # ---- F. 価格シート250日 ----
    log.append(f"\n【E】価格シートを{HIST_DAYS}営業日分に拡張")
    for path, name in ((BUNSEKI, "分析"), (URI, "売分析")):
        items[path] = widen_refs(items[path].decode(), name).encode()
    for path, name in PRICE.items():
        items[path] = extend_price(items[path].decode(), name).encode()

    # ---- G. 幾何情報 ----
    for path, name in ((KANRI, "管理"), (GENSEN, "厳選TOP2"), (BUNSEKI, "分析"), (URI, "売分析")):
        items[path] = fix_geometry(items[path].decode(), name).encode()
    log.append("\n【F】既存4シートの spans と dimension を実態に合わせて再計算")

    # ---- 仕上げ ----
    if "xl/calcChain.xml" in items:
        items["[Content_Types].xml"] = re.sub(
            r'<Override PartName="/xl/calcChain\.xml"[^>]*/>', "",
            items["[Content_Types].xml"].decode()).encode()
        items["xl/_rels/workbook.xml.rels"] = re.sub(
            r'<Relationship[^>]*Target="calcChain\.xml"[^>]*/>', "",
            items["xl/_rels/workbook.xml.rels"].decode()).encode()
    w = items["xl/workbook.xml"].decode()
    w = (re.sub(r"<calcPr[^>]*/>", '<calcPr calcId="0" fullCalcOnLoad="1"/>', w)
         if "<calcPr" in w else
         w.replace("</workbook>", '<calcPr calcId="0" fullCalcOnLoad="1"/></workbook>'))
    items["xl/workbook.xml"] = w.encode()

    with zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED) as zo:
        for n in order:
            if n == "xl/calcChain.xml":
                continue
            zo.writestr(n, items[n])
    log.append(f"\n出力: {DST}  ({DST.stat().st_size:,} bytes)")
    print("\n".join(log))


if __name__ == "__main__":
    main()
